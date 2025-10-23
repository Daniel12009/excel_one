# app.py (Versão 11 - Correção Venda por publicidade)
from flask import Flask, render_template, request, send_file, session
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
import io
import numpy as np
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
import re
import os
import traceback # Para log de erros detalhado

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024 * 5
TEMP_FILE_NAME = 'planilha_combinada_temp.xlsx'


# --- MAPEAMENTOS GLOBAIS ---

MAPA_CONTA = {
    'Decarion (Monaco Metais)': 'DECARION TORNEIRAS',
    'Gs Torneiras': 'GS TORNEIRAS',
    'Via Flix (Matriz)': 'VIA FLIX'
}

COLUNAS_GABARITO_FINAL = [
    'SKU PRINCIPAL', 'SKU', 'Data da venda', 'EMISSAO', 'N.º de venda',
    'origem', '# de anúncio', 'tipo de anuncio', 'Venda por publicidade',
    'Forma de entrega', 'Preço unitário de venda do anúncio (BRL)',
    'Unidades', 'Receita por produtos (BRL)', 'Envio Seller', 'TARIFA',
    'Tarifa de venda e impostos (BRL)', 'conta', 'Estado'
]

MAPA_ESTADOS = {
    'AC': 'Acre', 'AL': 'Alagoas', 'AP': 'Amapá', 'AM': 'Amazonas', 'BA': 'Bahia',
    'CE': 'Ceará', 'DF': 'Distrito Federal', 'ES': 'Espírito Santo', 'GO': 'Goiás',
    'MA': 'Maranhão', 'MT': 'Mato Grosso', 'MS': 'Mato Grosso do Sul', 'MG': 'Minas Gerais',
    'PA': 'Pará', 'PB': 'Paraíba', 'PR': 'Paraná', 'PE': 'Pernambuco', 'PI': 'Piauí',
    'RJ': 'Rio de Janeiro', 'RN': 'Rio Grande do Norte', 'RS': 'Rio Grande do Sul',
    'RO': 'Rondônia', 'RR': 'Roraima', 'SC': 'Santa Catarina', 'SP': 'São Paulo',
    'SE': 'Sergipe', 'TO': 'Tocantins'
}


# --- Funções Helper (Auxiliares) ---

def safe_decimal(value):
    """ (V6.2 - Robusta) """
    if pd.isna(value): return Decimal('0')
    if isinstance(value, Decimal): return value
    if isinstance(value, (int, float)): return Decimal(str(value))
    try:
        clean_value = str(value).strip().replace('R$', '').replace(' ', '')
        if '.' in clean_value and ',' in clean_value:
            clean_value = clean_value.replace('.', '')
            clean_value = clean_value.replace(',', '.')
        elif ',' in clean_value:
            clean_value = clean_value.replace(',', '.')
        if clean_value in ('', '-'): return Decimal('0')
        return Decimal(clean_value)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0')


def to_float_exact(value):
    """ Converte Decimal/str/int para float 2 casas """
    if not isinstance(value, Decimal):
        value = safe_decimal(value)
    quantized = value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    return float(quantized)

def formatar_data_br(value):
    """ Tenta formatar datas (DD/MM/YYYY ou YYYY-MM-DD) para DD/MM/YYYY """
    if pd.isna(value):
        return ''
    try:
        # Tenta o formato DD/MM/YYYY
        return pd.to_datetime(value, format='%d/%m/%Y').strftime('%d/%m/%Y')
    except ValueError:
        try:
            # Tenta formato padrão
            return pd.to_datetime(value).strftime('%d/%m/%Y')
        except Exception:
            # Tenta remover timezones incorretos comuns
            if isinstance(value, str):
                value_cleaned = re.sub(r'[+-]\d{2}:\d{2}$', '', value).strip()
                try:
                     return pd.to_datetime(value_cleaned).strftime('%d/%m/%Y')
                except Exception:
                    pass # Ignora se ainda falhar
            return str(value) # Retorna original se falhar


def format_excel_output(output_stream):
    """ (V8 - Atualizada para colunas vazias) """
    wb = load_workbook(output_stream)
    ws = wb.active
    alignment_center = Alignment(horizontal='center', vertical='center')
    header_font = Font(bold=True)
    format_brl = '#,##0.00'
    format_int = '0'
    brl_columns = ['Preço unitário de venda do anúncio (BRL)', 'Receita por produtos (BRL)', 'Envio Seller', 'TARIFA', 'Tarifa de venda e impostos (BRL)']
    int_columns = ['Unidades']
    brl_col_indices = []
    int_col_indices = []

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = alignment_center
        if cell.value in brl_columns:
            brl_col_indices.append(cell.column)
        if cell.value in int_columns:
            int_col_indices.append(cell.column)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = alignment_center
            if cell.column in brl_col_indices:
                cell.number_format = format_brl
            elif cell.column in int_col_indices:
                cell.number_format = format_int

    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                content = str(cell.value)
                if len(content) > max_length:
                    max_length = len(content)
            except: pass

        adjusted_width = (max_length + 2)
        if adjusted_width < 10: adjusted_width = 10
        if adjusted_width > 60: adjusted_width = 60
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.auto_filter.ref = ws.dimensions
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output

# --- Fim das Funções Helper ---


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/processar', methods=['POST'])
def processar_arquivos():

    all_processed_dfs = []
    cost_reference_dict = {}
    entry_index = 0

    try:
        while request.form.get(f'marketplace_{entry_index}') is not None:
            marketplace = request.form.get(f'marketplace_{entry_index}')
            conta_selecionada = request.form.get(f'conta_selecionada_{entry_index}')

            # Lógica de Conta Condicional (V9)
            if marketplace == 'SHOPEE/SHEIN' and conta_selecionada == 'Decarion (Monaco Metais)':
                conta_final = 'Via Flix - A Casa das Torneiras'
            else:
                conta_final = MAPA_CONTA.get(conta_selecionada, 'OUTRAS')

            files = request.files.getlist(f'files_{entry_index}[]')

            if not files or files[0].filename == '':
                entry_index += 1
                continue

            for file in files:
                if not file.filename.endswith('.xlsx'):
                    continue

                data_io = io.BytesIO(file.read())
                print(f"Processando arquivo: {file.filename} para {marketplace} / {conta_final}") # Log

                # --- LÓGICA DE ROTEAMENTO (V8) ---

                if marketplace == 'Mercado Livre':

                    # --- INÍCIO LÓGICA MERCADO LIVRE (V11) ---
                    df = pd.read_excel(data_io, header=5, dtype=str)
                    if df.empty: continue

                    # (Passo 1) Limpeza Inicial e Tipagem
                    df.rename(columns={
                        'Tipo de anúncio': 'tipo de anuncio', 'Cód. item': 'SKU',
                        'Cód. do item': 'SKU', 'SKU item': 'SKU',
                        'SKU da variação': 'SKU', 'Código': 'SKU'
                    }, inplace=True)

                    if 'Estado.1' not in df.columns: df['Estado.1'] = ''
                    # Garante que 'Venda por publicidade' exista
                    if 'Venda por publicidade' not in df.columns: df['Venda por publicidade'] = ''

                    text_cols = ['SKU', 'N.º de venda', 'Forma de entrega', 'Estado', 'Estado.1', 'tipo de anuncio', 'Venda por publicidade', '# de anúncio']
                    for col in text_cols:
                        if col not in df.columns: df[col] = ''
                        df[col] = df[col].fillna('').astype(str).str.strip()
                    df['N.º de venda'] = df['N.º de venda'].str.replace(r'\.0$', '', regex=True)

                    decimal_cols = ['Tarifas de envio (BRL)', 'Tarifa de venda e impostos (BRL)', 'TARIFA', 'Preço unitário de venda do anúncio (BRL)', 'Receita por produtos (BRL)']
                    for col in decimal_cols:
                        if col not in df.columns: df[col] = Decimal('0')
                        if col == 'Tarifa de venda e impostos (BRL)':
                             df[col] = Decimal('0')
                        else:
                             df[col] = df[col].apply(safe_decimal)

                    if 'Unidades' not in df.columns: df['Unidades'] = 0
                    df['Unidades'] = pd.to_numeric(df['Unidades'], errors='coerce').fillna(0).astype(int)

                    try:
                        month_map = { 'janeiro': '01', 'fevereiro': '02', 'março': '03', 'abril': '04', 'maio': '05', 'junho': '06', 'julho': '07', 'agosto': '08', 'setembro': '09', 'outubro': '10', 'novembro': '11', 'dezembro': '12' }
                        if 'Data da venda' in df.columns:
                            dates = df['Data da venda'].astype(str).str.replace(r'\s\d{1,2}:\d{2}\s*hs\.$', '', regex=True).str.strip()
                            for pt_month, num_month in month_map.items():
                                dates = dates.str.replace(f' de {pt_month} de ', f'/{num_month}/', regex=False)
                            df['Data da venda'] = pd.to_datetime(dates, format='%d/%m/%Y', errors='coerce').dt.strftime('%d/%m/%Y')
                            df['Data da venda'].fillna('', inplace=True)
                        else:
                             df['Data da venda'] = ''
                    except Exception as e:
                        print(f"Aviso (ML): Falha na formatação da 'Data da venda'. Erro: {e}")
                        if 'Data da venda' not in df.columns: df['Data da venda'] = ''
                        df['Data da venda'] = df['Data da venda'].apply(formatar_data_br) # Fallback


                    # (Passo 2) Propagação - SEM 'Venda por publicidade'
                    # --- CORREÇÃO (V11) ---
                    cols_to_propagate = ['Forma de entrega', 'tipo de anuncio', 'Estado.1']
                    # --- FIM CORREÇÃO ---
                    for col in cols_to_propagate:
                        if col in df.columns:
                            df[col] = df[col].replace(r'^\s*$', np.nan, regex=True)
                            df[col] = df[col].ffill()
                            df[col] = df[col].fillna('')

                    # --- NOVO (V11): Propagação Seletiva para 'Venda por publicidade' ---
                    # (Passo 2.5) Antes de deletar a linha master
                    if 'Venda por publicidade' in df.columns:
                         # Marca onde a linha atual é item (SKU preenchido) e a anterior era master (SKU vazio)
                         is_item_after_master = (df['SKU'].shift(1).fillna('') == '') & (df['SKU'] != '')
                         # Cria um índice para propagar o valor da master (shift(1))
                         master_index = df.index.to_series().shift(1)

                         # Condição para copiar:
                         # 1. Linha atual é um item que vem depois de uma master
                         # 2. 'Venda por publicidade' da linha atual está vazia
                         # 3. 'Venda por publicidade' da linha master (anterior) NÃO está vazia
                         condition_to_copy = (
                             is_item_after_master &
                             (df['Venda por publicidade'].replace('', np.nan).isna()) &
                             (df['Venda por publicidade'].shift(1).fillna('') != '')
                         )

                         # Aplica a cópia onde a condição é verdadeira
                         df.loc[condition_to_copy, 'Venda por publicidade'] = df['Venda por publicidade'].shift(1)[condition_to_copy]
                    # --- FIM NOVO (V11) ---

                    # (Passo 3) Máscara Resumo
                    summary_lines_mask = df['SKU'].str.strip() == ''

                    # (Passo 4) Deletar Linhas Indesejadas
                    df = df[~df['Estado'].astype(str).str.contains('cancel', case=False, na=False)].copy()
                    df = df[~summary_lines_mask].copy()
                    if df.empty: continue

                    df['Preço unitário de venda do anúncio (BRL)'] = df['Preço unitário de venda do anúncio (BRL)'].apply(safe_decimal)
                    df['Receita por produtos (BRL)'] = df['Unidades'].apply(Decimal) * df['Preço unitário de venda do anúncio (BRL)']

                    # (Passo 5) Herança (N e O)
                    df['Price_Key_Decimal'] = df['Preço unitário de venda do anúncio (BRL)'].apply(lambda x: x.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                    if cost_reference_dict:
                        ref_data = [(k[0], k[1], k[2], v[0], v[1]) for k, v in cost_reference_dict.items()]
                        ref_df = pd.DataFrame(ref_data, columns=['SKU', 'Price_Key_Decimal', 'Unidades', 'Ref_N', 'Ref_O'])
                        if not ref_df.empty:
                            df = df.merge(ref_df, on=['SKU', 'Price_Key_Decimal', 'Unidades'], how='left')
                            df['Ref_N'] = df['Ref_N'].fillna(np.nan)
                            df['Ref_O'] = df['Ref_O'].fillna(np.nan)
                            cond_heranca_n = pd.notna(df['Ref_N']) & (df['Ref_N'] > df['Tarifas de envio (BRL)'])
                            df['Tarifas de envio (BRL)'] = np.where(cond_heranca_n, df['Ref_N'], df['Tarifas de envio (BRL)'])
                            cond_heranca_o = pd.notna(df['Ref_O']) & (df['Ref_O'] > df['TARIFA'])
                            df['TARIFA'] = np.where(cond_heranca_o, df['Ref_O'], df['TARIFA'])
                            df.drop(columns=['Ref_N', 'Ref_O'], inplace=True, errors='ignore')

                    # (Passo 6) Regra < 79 (N)
                    cond_lt_79 = df['Preço unitário de venda do anúncio (BRL)'] < Decimal('79.00')
                    df['Tarifas de envio (BRL)'] = np.where(cond_lt_79, Decimal('0'), df['Tarifas de envio (BRL)'])

                    # (Passo 7) Regra Flex (N) (V10)
                    cond_flex = df['Forma de entrega'] == 'Mercado Envios Flex'
                    cond_flex_high = cond_flex & (df['Preço unitário de venda do anúncio (BRL)'] >= Decimal('79.00'))
                    cond_flex_low = cond_flex & (df['Preço unitário de venda do anúncio (BRL)'] < Decimal('79.00'))
                    df['Tarifas de envio (BRL)'] = np.select(
                        [cond_flex_high, cond_flex_low],
                        [Decimal('-9.11'), Decimal('-1.10')],
                        default=df['Tarifas de envio (BRL)']
                    )

                    # (Passo 7.5) Cálculo P
                    preco_unit = df['Preço unitário de venda do anúncio (BRL)']
                    cond_fixo_1 = (preco_unit > Decimal('0')) & (preco_unit <= Decimal('12.50'))
                    val_fixo_1 = preco_unit * Decimal('-0.5')
                    cond_fixo_2 = (preco_unit > Decimal('12.50')) & (preco_unit <= Decimal('29.00'))
                    val_fixo_2 = Decimal('-6.25')
                    cond_fixo_3 = (preco_unit > Decimal('29.00')) & (preco_unit <= Decimal('50.00'))
                    val_fixo_3 = Decimal('-6.50')
                    cond_fixo_4 = (preco_unit > Decimal('50.00')) & (preco_unit < Decimal('79.00'))
                    val_fixo_4 = Decimal('-6.75')
                    custo_fixo_unitario = np.select(
                        [cond_fixo_1, cond_fixo_2, cond_fixo_3, cond_fixo_4],
                        [val_fixo_1, val_fixo_2, val_fixo_3, val_fixo_4],
                        default=Decimal('0')
                    )
                    unidades_decimal = df['Unidades'].apply(Decimal)
                    custo_fixo_total = custo_fixo_unitario * unidades_decimal
                    receita = df['Receita por produtos (BRL)']
                    custo_11_5 = receita * Decimal('-0.115')
                    df['Tarifa de venda e impostos (BRL)'] = (custo_11_5 + custo_fixo_total).apply(
                         lambda x: x.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    )

                    # (Passo 9) Estado Final
                    if 'Estado.1' in df.columns:
                        df['Estado'] = df['Estado.1']
                    df.drop(columns=['Estado.1', 'Price_Key_Decimal'], inplace=True, errors='ignore')

                    # (Passo 10) Finalização
                    df['SKU PRINCIPAL'] = df['SKU']
                    df['Envio Seller'] = df['Tarifas de envio (BRL)']
                    df['EMISSAO'] = df['Data da venda']
                    df['origem'] = marketplace
                    df['conta'] = conta_final

                    # Atualização da Referência (N e O)
                    df_for_ref = df.copy()
                    df_for_ref['Preço unitário de venda do anúncio (BRL)'] = df_for_ref['Preço unitário de venda do anúncio (BRL)'].apply(safe_decimal)
                    df_for_ref['Price_Key_Decimal'] = df_for_ref['Preço unitário de venda do anúncio (BRL)'].apply(lambda x: x.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                    df_for_ref['Ref_N'] = df_for_ref['Envio Seller'].apply(safe_decimal)
                    df_for_ref['Ref_O'] = df_for_ref['TARIFA'].apply(safe_decimal)
                    for _, row in df_for_ref.iterrows():
                        key = (row['SKU'], row['Price_Key_Decimal'], row['Unidades'])
                        new_costs = (row['Ref_N'], row['Ref_O'])
                        if key not in cost_reference_dict:
                            cost_reference_dict[key] = new_costs
                        else:
                            current_costs = cost_reference_dict[key]
                            best_n = max(current_costs[0], new_costs[0])
                            best_o = max(current_costs[1], new_costs[1])
                            cost_reference_dict[key] = (best_n, best_o)

                    # --- FIM LÓGICA MERCADO LIVRE ---


                elif marketplace == 'SHOPEE/SHEIN':

                    # --- INÍCIO LÓGICA SHOPEE/SHEIN (V10) ---

                    # (Passo 0) Leitura
                    df = pd.read_excel(data_io, header=0, dtype=str)
                    if df.empty: continue

                    # (Passo 1) Renomeação e Limpeza
                    df.rename(columns={
                        'Código (SKU)': 'SKU',
                        'Número do pedido no e-commerce': 'N.º de venda',
                        'E-commerce': 'origem',
                        'UF': 'Estado'
                    }, inplace=True)

                    # (Passo 2) Filtro de Cancelados
                    if 'Situação da venda' in df.columns:
                        df = df[~df['Situação da venda'].astype(str).str.contains('cancel', case=False, na=False)].copy()
                    if df.empty: continue

                    # (Passo 3) Mapeamento para Gabarito Final
                    df_final = pd.DataFrame()

                    df_final['SKU PRINCIPAL'] = df['SKU']
                    df_final['SKU'] = df['SKU']
                    df_final['N.º de venda'] = df['N.º de venda']
                    df_final['origem'] = df['origem']
                    df_final['conta'] = conta_final
                    df_final['Estado'] = df['Estado'].str.upper().map(MAPA_ESTADOS).fillna(df['Estado'])

                    df_final['Data da venda'] = df['Data da venda'].apply(formatar_data_br)
                    df_final['EMISSAO'] = df['Data de Emissão'].apply(formatar_data_br)

                    df_final['# de anúncio'] = ''
                    df_final['tipo de anuncio'] = ''
                    df_final['Venda por publicidade'] = '' # Tiny não tem essa info, fica em branco
                    df_final['Forma de entrega'] = ''

                    df_final['Unidades'] = pd.to_numeric(df['Quantidade de produtos'], errors='coerce').fillna(0).astype(int)

                    preco_unit = df['Preço unitário'].apply(safe_decimal)
                    preco_total = df['Preço total'].apply(safe_decimal)
                    desconto = df['Valor de desconto'].apply(safe_decimal)
                    comissao = df['Comissão e-commerce'].apply(safe_decimal)

                    # Lógica de Custo Shein (V10)
                    cond_shein = df['origem'] == 'Shein'
                    envio_seller_tiny = np.where(cond_shein, Decimal('-6.00'), Decimal('0'))

                    df_final['Preço unitário de venda do anúncio (BRL)'] = preco_unit
                    df_final['Receita por produtos (BRL)'] = preco_total - desconto
                    df_final['Envio Seller'] = envio_seller_tiny
                    df_final['TARIFA'] = Decimal('0')
                    df_final['Tarifa de venda e impostos (BRL)'] = comissao * Decimal('-1')

                    df = df_final
                    # --- FIM LÓGICA SHOPEE/SHEIN ---


                # --- PÓS-PROCESSAMENTO (Comum a ambos) ---

                # Converte para Float (para o Excel)
                float_cols = ['Envio Seller', 'TARIFA', 'Tarifa de venda e impostos (BRL)', 'Preço unitário de venda do anúncio (BRL)', 'Receita por produtos (BRL)']
                for col in float_cols:
                    if col in df.columns:
                        # Aplica safe_decimal antes de to_float_exact por segurança
                        df[col] = df[col].apply(safe_decimal).apply(to_float_exact)
                df['Unidades'] = df['Unidades'].astype(int)

                # Garante colunas do gabarito
                for col in COLUNAS_GABARITO_FINAL:
                    if col not in df.columns:
                        df[col] = ''

                df_final_file = df.reindex(columns=COLUNAS_GABARITO_FINAL).fillna('')
                all_processed_dfs.append(df_final_file)

            entry_index += 1

        # 4. CONCATENAÇÃO E ORDENAÇÃO
        if not all_processed_dfs:
            return "Nenhum arquivo válido foi processado.", 400

        final_df = pd.concat(all_processed_dfs, ignore_index=True)
        final_df = final_df.sort_values(by=['conta', 'SKU PRINCIPAL'], ascending=[True, True])

        # 5. GERAÇÃO E FORMATAÇÃO DO EXCEL
        output_stream = io.BytesIO()
        writer = pd.ExcelWriter(output_stream, engine='openpyxl')
        final_df.to_excel(writer, index=False, sheet_name='Planilha_Combinada')
        writer.close()
        output_stream.seek(0)

        # Aplica formatação openpyxl
        final_output_stream = format_excel_output(output_stream)

        # 7. Resposta
        return send_file(
            final_output_stream,
            as_attachment=True,
            download_name='planilha_combinada_formatada.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(traceback.format_exc()) # Log completo do erro no console
        return f"<h1>Erro no Processamento</h1><p>Ocorreu um erro inesperado. Detalhes: <b>{e}</b></p><p>Verifique o formato dos arquivos e a lógica de processamento.</p>", 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
